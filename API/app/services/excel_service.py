import base64
import io
import os
import tempfile
from datetime import date
from decimal import Decimal

import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.enums import STATUT_AFFAIRE_LABELS
from app.models.affaire import Affaire
from app.models.demande import Demande
from app.models.point_traitement import FichePointTraitement
from app.schemas.action import PlanActionLigne

def _fmt_date(d) -> str:
    if not d:
        return ""
    if isinstance(d, date):
        return d.strftime("%d/%m/%Y")
    return str(d)

def _fmt_montant(m) -> str:
    if m is None:
        return ""
    try:
        return f"{m:,.0f}".replace(",", " ") + " FCFA"
    except Exception:
        return str(m)

def build_fiche_affaire_excel(affaire: Affaire) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Fiche_{affaire.numero_affaire}"
    
    # Visible grid lines
    ws.views.sheetView[0].showGridLines = True
    
    # Design style definitions
    font_family = "Segoe UI"
    primary_green_fill = PatternFill(start_color="008940", end_color="008940", fill_type="solid")
    light_green_fill = PatternFill(start_color="E8F4EE", end_color="E8F4EE", fill_type="solid")
    
    font_title = Font(name=font_family, size=16, bold=True, color="FFFFFF")
    font_subtitle = Font(name=font_family, size=9, italic=True, color="555555")
    font_section = Font(name=font_family, size=12, bold=True, color="008940")
    font_header = Font(name=font_family, size=10, bold=True, color="FFFFFF")
    font_bold_dark = Font(name=font_family, size=10, bold=True, color="000000")
    font_regular = Font(name=font_family, size=10, color="000000")
    
    thin_border_side = Side(border_style="thin", color="D3D3D3")
    thin_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center")
    
    # 1. Header
    ws.cell(row=1, column=1, value="ISITEK SARL — Depuis 2012, votre partenaire technique de confiance").font = font_subtitle
    ws.merge_cells("A3:D3")
    title_cell = ws.cell(row=3, column=1, value="FICHE D'AFFAIRE")
    title_cell.font = font_title
    title_cell.fill = primary_green_fill
    title_cell.alignment = align_center
    ws.row_dimensions[3].height = 40
    
    # 2. General Info Table
    info_fields = [
        ("N° AFFAIRE", affaire.numero_affaire, "RESPONSABLE", f"{affaire.responsable_prenom} {affaire.responsable_nom}"),
        ("DATE D'OUVERTURE", _fmt_date(affaire.date_ouverture), "CLIENT", affaire.client_nom),
        ("N° COMMANDE", affaire.numero_commande or "", "DATE LIVRAISON BC", _fmt_date(affaire.date_livraison_bc))
    ]
    
    current_row = 5
    for labels_vals in info_fields:
        c = ws.cell(row=current_row, column=1, value=labels_vals[0])
        c.font = font_bold_dark
        c.fill = light_green_fill
        c.border = thin_border
        
        c = ws.cell(row=current_row, column=2, value=labels_vals[1])
        c.font = font_regular
        c.border = thin_border
        
        c = ws.cell(row=current_row, column=3, value=labels_vals[2])
        c.font = font_bold_dark
        c.fill = light_green_fill
        c.border = thin_border
        
        c = ws.cell(row=current_row, column=4, value=labels_vals[3])
        c.font = font_regular
        c.border = thin_border
        
        ws.row_dimensions[current_row].height = 20
        current_row += 1
        
    current_row += 1 # Empty spacing
    
    # 3. Extra Info Table
    extra_fields = [
        ("LIBELLÉ AFFAIRE", affaire.libelle_affaire),
        ("DOMAINE", affaire.domaine),
        ("TYPE AFFAIRE", affaire.type_affaire or ""),
        ("CORRESPONDANT", f"{affaire.correspondant_nom or ''} — {affaire.correspondant_telephone or ''} — {affaire.correspondant_email or ''}"),
        ("MONTANT", _fmt_montant(affaire.montant_affaire)),
        ("STATUT", STATUT_AFFAIRE_LABELS.get(affaire.statut.value, affaire.statut.value))
    ]
    
    for label, val in extra_fields:
        c = ws.cell(row=current_row, column=1, value=label)
        c.font = font_bold_dark
        c.fill = light_green_fill
        c.border = thin_border
        
        ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=4)
        for col in range(2, 5):
            ws.cell(row=current_row, column=col).border = thin_border
            
        c_val = ws.cell(row=current_row, column=2, value=val)
        c_val.font = font_regular
        
        ws.row_dimensions[current_row].height = 20
        current_row += 1
        
    current_row += 1 # Empty spacing
    
    # 4. Progression Affaire Table
    ws.cell(row=current_row, column=1, value="PROGRESSION AFFAIRE").font = font_section
    ws.row_dimensions[current_row].height = 24
    current_row += 1
    
    headers = ["ÉTAPE / ACTION", "DÉBUT / ACTION", "FIN / REF / DETAIL", "OBSERVATIONS / COMMENTAIRES"]
    for col_idx, h in enumerate(headers, 1):
        c = ws.cell(row=current_row, column=col_idx, value=h)
        c.font = font_header
        c.fill = primary_green_fill
        c.border = thin_border
        c.alignment = align_center
    ws.row_dimensions[current_row].height = 25
    current_row += 1
    
    for action in affaire.actions:
        col1 = _fmt_date(action.date_debut or action.date_action)
        col2 = _fmt_date(action.date_fin) or (action.ref or "")
        if action.fournisseur:
            col2 = action.fournisseur
        if action.mode:
            col2 = action.mode
        if action.agence:
            col2 = action.agence
        if action.banque:
            col2 = action.banque.nom
            
        obs = action.observations or action.commentaire or ""
        check = "☑  " if action.termine else "☐  "
        
        c = ws.cell(row=current_row, column=1, value=f"{check}{action.libelle}")
        c.font = font_bold_dark if action.termine else font_regular
        c.border = thin_border
        c.alignment = align_left
        
        c = ws.cell(row=current_row, column=2, value=col1)
        c.font = font_regular
        c.border = thin_border
        c.alignment = align_center
        
        c = ws.cell(row=current_row, column=3, value=col2)
        c.font = font_regular
        c.border = thin_border
        c.alignment = align_center
        
        c = ws.cell(row=current_row, column=4, value=obs)
        c.font = font_regular
        c.border = thin_border
        c.alignment = align_left
        
        ws.row_dimensions[current_row].height = 22
        current_row += 1
        
    current_row += 2 # Empty spacing
    
    # 5. Signatures Grid
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=4)
    sig_title = ws.cell(row=current_row, column=1, value="SIGNATURES ET VALIDATION")
    sig_title.font = font_bold_dark
    sig_title.alignment = align_center
    current_row += 1
    
    sig_headers = ["Responsable affaire", "Comptable", "Business Controller", "Direction"]
    for col_idx, h in enumerate(sig_headers, 1):
        c = ws.cell(row=current_row, column=col_idx, value=h)
        c.font = font_header
        c.fill = primary_green_fill
        c.border = thin_border
        c.alignment = align_center
    ws.row_dimensions[current_row].height = 20
    current_row += 1
    
    # Box for signatures
    ws.row_dimensions[current_row].height = 45
    for col_idx in range(1, 5):
        c = ws.cell(row=current_row, column=col_idx)
        c.border = thin_border
        
    # Auto-adjust column widths
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 35
    
    output = io.BytesIO()
    wb.save(output)
    wb.close()
    return output.getvalue()

def build_plan_action_excel(lignes: list[PlanActionLigne], titre: str = "PLAN D'ACTIONS") -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Plan_Actions"
    
    ws.views.sheetView[0].showGridLines = True
    
    font_family = "Segoe UI"
    primary_green_fill = PatternFill(start_color="008940", end_color="008940", fill_type="solid")
    light_green_fill = PatternFill(start_color="E8F4EE", end_color="E8F4EE", fill_type="solid")
    
    font_title = Font(name=font_family, size=14, bold=True, color="FFFFFF")
    font_header = Font(name=font_family, size=9, bold=True, color="FFFFFF")
    font_regular = Font(name=font_family, size=9, color="000000")
    font_bold = Font(name=font_family, size=9, bold=True, color="000000")
    font_subtitle = Font(name=font_family, size=9, italic=True, color="555555")
    
    thin_border_side = Side(border_style="thin", color="D3D3D3")
    thin_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center")
    
    ws.merge_cells("A1:I1")
    title_cell = ws.cell(row=1, column=1, value=titre.upper())
    title_cell.font = font_title
    title_cell.fill = primary_green_fill
    title_cell.alignment = align_center
    ws.row_dimensions[1].height = 35
    
    ws.cell(row=2, column=1, value="Priorité : Haute | Moyenne | Basse").font = font_subtitle
    ws.row_dimensions[2].height = 18
    
    headers = [
        "N°",
        "ACTION / TÂCHES",
        "CLIENT",
        "RESPONSABLE",
        "SUPPORT",
        "DÉBUT",
        "FIN",
        "STATUT",
        "COMMENTAIRES",
    ]
    
    for col_idx, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=col_idx, value=h)
        c.font = font_header
        c.fill = primary_green_fill
        c.border = thin_border
        c.alignment = align_center
    ws.row_dimensions[4].height = 25
    
    current_row = 5
    for ligne in lignes:
        fill_to_use = light_green_fill if current_row % 2 == 0 else PatternFill(fill_type=None)
        
        vals = [
            str(ligne.numero),
            ligne.action,
            ligne.client,
            ligne.responsable,
            ligne.support,
            ligne.debut,
            ligne.fin,
            ligne.statut,
            ligne.commentaire
        ]
        
        for col_idx, val in enumerate(vals, 1):
            c = ws.cell(row=current_row, column=col_idx, value=val)
            c.font = font_regular
            c.border = thin_border
            if fill_to_use.fill_type:
                c.fill = fill_to_use
            
            if col_idx in [1, 6, 7, 8]:
                c.alignment = align_center
            else:
                c.alignment = align_left
                
        ws.row_dimensions[current_row].height = 20
        current_row += 1
        
    current_row += 2
    
    ws.cell(row=current_row, column=1, value="Validation Direction : _______________________").font = font_bold
    
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 12
    ws.column_dimensions["H"].width = 15
    ws.column_dimensions["I"].width = 25
    
    output = io.BytesIO()
    wb.save(output)
    wb.close()
    return output.getvalue()


def build_facture_excel(demande: Demande) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Facture_{demande.id}"
    
    ws.views.sheetView[0].showGridLines = True
    
    font_family = "Segoe UI"
    primary_green_fill = PatternFill(start_color="008940", end_color="008940", fill_type="solid")
    light_green_fill = PatternFill(start_color="E8F4EE", end_color="E8F4EE", fill_type="solid")
    
    font_title = Font(name=font_family, size=16, bold=True, color="FFFFFF")
    font_section = Font(name=font_family, size=11, bold=True, color="008940")
    font_header = Font(name=font_family, size=10, bold=True, color="FFFFFF")
    font_bold_dark = Font(name=font_family, size=10, bold=True, color="000000")
    font_regular = Font(name=font_family, size=10, color="000000")
    font_small = Font(name=font_family, size=9, color="555555")
    font_italic = Font(name=font_family, size=9, italic=True, color="777777")
    
    thin_border_side = Side(border_style="thin", color="D3D3D3")
    thin_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    
    # 1. Header
    ws.merge_cells("A1:E1")
    title_cell = ws.cell(row=1, column=1, value="FACTURE DE PRESTATION DE SERVICE")
    title_cell.font = font_title
    title_cell.fill = primary_green_fill
    title_cell.alignment = align_center
    ws.row_dimensions[1].height = 40
    
    # Company info on left, Invoice info on right
    ws.cell(row=3, column=1, value="EMETTEUR :").font = font_bold_dark
    ws.cell(row=4, column=1, value="ISITEK SARL").font = font_bold_dark
    ws.cell(row=5, column=1, value="Cocody Angré, Abidjan").font = font_small
    ws.cell(row=6, column=1, value="Téléphone : +225 07 48 00 00 00").font = font_small
    ws.cell(row=7, column=1, value="Email : contact@isitek.ci").font = font_small
    
    ws.cell(row=3, column=4, value="FACTURE N° :").font = font_bold_dark
    ws.cell(row=3, column=5, value=f"FAC-FNE-{demande.id:04d}").font = font_bold_dark
    ws.cell(row=4, column=4, value="DATE :").font = font_regular
    ws.cell(row=4, column=5, value=_fmt_date(demande.created_at or date.today())).font = font_regular
    ws.cell(row=5, column=4, value="STATUT :").font = font_regular
    ws.cell(row=5, column=5, value="PAYÉ" if demande.statut == "termine" else "À PAYER").font = font_bold_dark
    
    # Client info
    ws.cell(row=9, column=1, value="FACTURÉ À :").font = font_section
    client_name = f"{demande.client.prenom} {demande.client.nom}" if demande.client else "Client Inconnu"
    ws.cell(row=10, column=1, value=client_name).font = font_bold_dark
    ws.cell(row=11, column=1, value=f"Adresse : {demande.adresse or ''}").font = font_regular
    if demande.client and getattr(demande.client, "telephone", None):
        ws.cell(row=12, column=1, value=f"Tél : {demande.client.telephone}").font = font_regular
    
    # Details table
    ws.cell(row=15, column=1, value="DÉSIGNATION DES PRESTATIONS").font = font_section
    ws.row_dimensions[15].height = 20
    
    headers = ["DOMAINE", "TYPE PRESTATION", "DESCRIPTION", "MONTANT"]
    ws.merge_cells("C16:D16")
    
    # Set headers
    for i, h in enumerate(headers):
        c = ws.cell(row=16, column=1 + (i if i < 3 else 4), value=h)
        c.font = font_header
        c.fill = primary_green_fill
        c.border = thin_border
        c.alignment = align_center
    
    # Format the merged cell border for C16:D16
    ws.cell(row=16, column=3).border = thin_border
    ws.cell(row=16, column=4).border = thin_border
    
    ws.row_dimensions[16].height = 24
    
    # Data row
    row_idx = 17
    ws.merge_cells("C17:D17")
    c1 = ws.cell(row=row_idx, column=1, value=demande.domaine)
    c1.font = font_regular
    c1.border = thin_border
    c1.alignment = align_left
    
    c2 = ws.cell(row=row_idx, column=2, value=demande.type_prestation)
    c2.font = font_regular
    c2.border = thin_border
    c2.alignment = align_left
    
    c3 = ws.cell(row=row_idx, column=3, value=demande.description)
    c3.font = font_regular
    c3.border = thin_border
    c3.alignment = align_left
    ws.cell(row=row_idx, column=4).border = thin_border  # part of merge
    
    montant_val = demande.devis_montant if demande.devis_montant is not None else 0
    c4 = ws.cell(row=row_idx, column=5, value=montant_val)
    c4.number_format = '#,##0" FCFA"'
    c4.font = font_bold_dark
    c4.border = thin_border
    c4.alignment = align_right
    
    ws.row_dimensions[row_idx].height = 40
    
    # Totals
    row_idx += 2
    ws.cell(row=row_idx, column=4, value="TOTAL H.T.").font = font_bold_dark
    ws.cell(row=row_idx, column=4).alignment = align_right
    c_ht = ws.cell(row=row_idx, column=5, value=montant_val)
    c_ht.number_format = '#,##0" FCFA"'
    c_ht.font = font_regular
    c_ht.alignment = align_right
    c_ht.border = thin_border
    
    row_idx += 1
    ws.cell(row=row_idx, column=4, value="TVA (0%)").font = font_bold_dark
    ws.cell(row=row_idx, column=4).alignment = align_right
    c_tva = ws.cell(row=row_idx, column=5, value=0)
    c_tva.number_format = '#,##0" FCFA"'
    c_tva.font = font_regular
    c_tva.alignment = align_right
    c_tva.border = thin_border
    
    row_idx += 1
    ws.cell(row=row_idx, column=4, value="NET À PAYER").font = font_header
    ws.cell(row=row_idx, column=4).fill = primary_green_fill
    ws.cell(row=row_idx, column=4).alignment = align_right
    c_net = ws.cell(row=row_idx, column=5, value=montant_val)
    c_net.number_format = '#,##0" FCFA"'
    c_net.font = font_header
    c_net.fill = primary_green_fill
    c_net.alignment = align_right
    c_net.border = thin_border
    
    # Bank coordinates
    row_idx += 3
    ws.cell(row=row_idx, column=1, value="COORDONNÉES BANCAIRES POUR RÈGLEMENT").font = font_section
    ws.row_dimensions[row_idx].height = 20
    
    row_idx += 1
    bank_info = [
        ("Banque", "BICICI Côte d'Ivoire"),
        ("Code Banque", "CI006"),
        ("Code Guichet", "01001"),
        ("N° de Compte", "012345678901"),
        ("Clé RIB", "45"),
        ("RIB complet", "CI93 CI00 6010 0101 2345 6789 0145")
    ]
    
    for label, val in bank_info:
        c_lbl = ws.cell(row=row_idx, column=1, value=label)
        c_lbl.font = font_bold_dark
        c_lbl.fill = light_green_fill
        c_lbl.border = thin_border
        
        ws.merge_cells(start_row=row_idx, start_column=2, end_row=row_idx, end_column=3)
        ws.cell(row=row_idx, column=2).border = thin_border
        ws.cell(row=row_idx, column=3).border = thin_border
        c_val = ws.cell(row=row_idx, column=2, value=val)
        c_val.font = font_regular
        
        ws.row_dimensions[row_idx].height = 18
        row_idx += 1
        
    # Note / Footer
    row_idx += 2
    ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=5)
    c_foot = ws.cell(row=row_idx, column=1, value="Nous vous remercions pour votre confiance et votre fidélité.")
    c_foot.font = font_italic
    c_foot.alignment = align_center
    
    # Set columns widths
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 22
    
    output = io.BytesIO()
    wb.save(output)
    wb.close()
    return output.getvalue()


def _add_image_to_sheet(ws, image_path: str, anchor: str, width: int = 90, height: int = 90):
    if not os.path.exists(image_path):
        return
    img = XLImage(image_path)
    img.width = width
    img.height = height
    ws.add_image(img, anchor)


def _save_base64_image(base64_data: str) -> str | None:
    if not base64_data:
        return None
    try:
        raw = base64_data
        if "," in raw:
            raw = raw.split(",", 1)[1]
        data = base64.b64decode(raw)
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".png")
        tmp.write(data)
        tmp.close()
        return tmp.name
    except Exception:
        return None


def build_point_traitement_excel(fiche: FichePointTraitement) -> bytes:
    """Génère la fiche POINT TRAITEMENT DES DEMANDES identique au formulaire papier."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Point_Traitement"

    ws.views.sheetView[0].showGridLines = False

    font_family = "Arial"
    grey_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    black_side = Side(border_style="thin", color="000000")
    black_border = Border(left=black_side, right=black_side, top=black_side, bottom=black_side)

    font_title = Font(name=font_family, size=14, bold=True, color="000000")
    font_header = Font(name=font_family, size=9, bold=True, color="000000")
    font_bold = Font(name=font_family, size=10, bold=True, color="000000")
    font_regular = Font(name=font_family, size=9, color="000000")
    font_underline = Font(name=font_family, size=10, bold=True, underline="single", color="000000")

    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    align_right = Alignment(horizontal="right", vertical="center")

    col_widths = {"A": 4, "B": 11, "C": 16, "D": 13, "E": 32, "F": 12, "G": 13, "H": 12}
    for col, w in col_widths.items():
        ws.column_dimensions[col].width = w

    logo_path = os.path.normpath(
        os.path.join(os.path.dirname(__file__), "..", "..", "assets", "logo_isitek.png")
    )
    ws.merge_cells("A1:A4")
    ws.row_dimensions[1].height = 18
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 18
    ws.row_dimensions[4].height = 18
    _add_image_to_sheet(ws, logo_path, "A1", width=70, height=70)

    ws.merge_cells("B1:H2")
    title_cell = ws["B1"]
    title_cell.value = "POINT TRAITEMENT DES DEMANDES"
    title_cell.font = font_title
    title_cell.fill = grey_fill
    title_cell.alignment = align_center
    title_cell.border = black_border
    for col in range(2, 9):
        c = ws.cell(row=1, column=col)
        c.fill = grey_fill
        c.border = black_border
        ws.cell(row=2, column=col).fill = grey_fill
        ws.cell(row=2, column=col).border = black_border

    ws.merge_cells("B3:H3")
    semaine_txt = f"SEMAINE {fiche.semaine or '........'}"
    debut_txt = _fmt_date(fiche.date_debut) if fiche.date_debut else "................"
    fin_txt = _fmt_date(fiche.date_fin) if fiche.date_fin else "................"
    semaine_cell = ws["B3"]
    semaine_cell.value = f"{semaine_txt}          DU {debut_txt}          AU {fin_txt}"
    semaine_cell.font = font_bold
    semaine_cell.alignment = align_left
    semaine_cell.border = black_border

    ws.merge_cells("B4:H4")
    resp_cell = ws["B4"]
    resp_cell.value = f"RESPONSABLE : {fiche.responsable or ''}"
    resp_cell.font = font_bold
    resp_cell.alignment = align_left
    resp_cell.border = black_border

    header_row = 5
    headers = ["N°", "DATE", "CLIENT", "REF DEMANDE", "RESUME DEMANDE", "REF DEVIS", "MONTANT HT", "STATUT"]
    for col_idx, h in enumerate(headers, 1):
        c = ws.cell(row=header_row, column=col_idx, value=h)
        c.font = font_header
        c.border = black_border
        c.alignment = align_center
    ws.row_dimensions[header_row].height = 22

    lignes_by_num = {l.numero: l for l in fiche.lignes}
    data_start = 6
    montant_rows = []
    for i in range(1, 11):
        row = data_start + i - 1
        ligne = lignes_by_num.get(i)
        ws.cell(row=row, column=1, value=i).font = font_regular
        ws.cell(row=row, column=1).border = black_border
        ws.cell(row=row, column=1).alignment = align_center

        date_val = _fmt_date(ligne.date_demande) if ligne and ligne.date_demande else ""
        client_val = (ligne.client or "") if ligne else ""
        ref_dem_val = (ligne.ref_demande or "") if ligne else ""
        resume_val = (ligne.resume_demande or "") if ligne else ""
        ref_dev_val = (ligne.ref_devis or "") if ligne else ""
        montant_val = ligne.montant_ht if ligne and ligne.montant_ht is not None else None
        statut_val = (ligne.statut or "") if ligne else ""

        vals = [date_val, client_val, ref_dem_val, resume_val, ref_dev_val, montant_val, statut_val]
        for col_offset, val in enumerate(vals, 2):
            c = ws.cell(row=row, column=col_offset, value=val)
            c.font = font_regular
            c.border = black_border
            if col_offset == 7 and montant_val is not None:
                c.number_format = '#,##0'
                c.alignment = align_right
                montant_rows.append(row)
            elif col_offset == 5:
                c.alignment = align_left
            else:
                c.alignment = align_center if col_offset in (2, 6, 8) else align_left

        ws.row_dimensions[row].height = 28

    footer_row = data_start + 11
    ws.row_dimensions[footer_row].height = 30
    ws.row_dimensions[footer_row + 1].height = 50

    ws.merge_cells(start_row=footer_row, start_column=1, end_row=footer_row + 1, end_column=4)
    sig_label = ws.cell(row=footer_row, column=1, value="Signature responsable")
    sig_label.font = font_underline
    sig_label.alignment = Alignment(horizontal="left", vertical="top")

    sig_temp = None
    if fiche.signature_base64:
        sig_temp = _save_base64_image(fiche.signature_base64)
        if sig_temp:
            _add_image_to_sheet(ws, sig_temp, f"B{footer_row + 1}", width=120, height=40)

    total_label_row = footer_row
    ws.cell(row=total_label_row, column=6, value="TOTAL").font = font_bold
    ws.cell(row=total_label_row, column=6).alignment = align_right

    ws.merge_cells(start_row=total_label_row, start_column=7, end_row=total_label_row, end_column=8)
    total_cell = ws.cell(row=total_label_row, column=7)
    if montant_rows:
        refs = "+".join(f"G{r}" for r in montant_rows)
        total_cell.value = f"={refs}"
    else:
        total_cell.value = 0
    total_cell.font = font_bold
    total_cell.border = black_border
    total_cell.alignment = align_center
    total_cell.number_format = '#,##0'
    ws.cell(row=total_label_row, column=8).border = black_border

    output = io.BytesIO()
    wb.save(output)
    wb.close()

    if sig_temp and os.path.exists(sig_temp):
        try:
            os.unlink(sig_temp)
        except OSError:
            pass

    return output.getvalue()
