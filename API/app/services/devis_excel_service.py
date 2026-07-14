"""Génération devis proforma ISITEK via Excel (modèle az.jpeg / ze.jpeg / er.jpeg)."""

from __future__ import annotations

import io
import os
import subprocess
import tempfile
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.models.devis_proforma import DevisProforma
from app.schemas.devis import DevisRenderRequest, ProduitLigne

ASSETS_DIR = Path(__file__).resolve().parents[2] / "assets" / "devis"
TEMPLATE_PATH = ASSETS_DIR / "devis_proforma_template.xlsx"

GREY = "D8D8D8"
RED = "CC0000"
BLACK = "000000"
THIN = Side(style="thin", color=BLACK)
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
THICK_TOP = Border(left=THIN, right=THIN, top=Side(style="medium", color=BLACK), bottom=THIN)

FONT = "Arial"
ITEM_ROWS = 12
ROW_HEADER = 10
ROW_ITEMS_START = 11
ROW_TOTALS_START = ROW_ITEMS_START + ITEM_ROWS


def _asset(name: str) -> Path | None:
    p = ASSETS_DIR / name
    return p if p.is_file() else None


def _fmt_num(n: float | int | None) -> str:
    if n is None or n == 0:
        return ""
    return f"{int(round(float(n))):,}".replace(",", " ")


def _compute_totals(lignes: list[ProduitLigne], rx_on: bool, rx_pct: float) -> tuple[int, int, int, int, int]:
    brut = remise = 0.0
    for l in lignes:
        line_brut = l.quantite * l.prix_unitaire_ht
        line_rem = line_brut * (l.remise_pourcentage / 100)
        brut += line_brut
        remise += line_rem
    sous = brut - remise
    rx = sous * (rx_pct / 100) if rx_on else 0
    net = sous - rx
    return int(round(brut)), int(round(remise)), int(round(sous)), int(round(rx)), int(round(net))


def _set_cell(ws, row: int, col: int, value="", *, bold=False, italic=False, red=False, size=9, align="left", fill=None, border=BORDER, wrap=False):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name=FONT, size=size, bold=bold, italic=italic, color=RED if red else BLACK)
    c.alignment = Alignment(
        horizontal={"left": "left", "right": "right", "center": "center"}.get(align, "left"),
        vertical="center",
        wrap_text=wrap,
    )
    if fill:
        c.fill = fill
    if border:
        c.border = border
    return c


def _merge_fill(ws, r1, c1, r2, c2, value="", **kwargs):
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    _set_cell(ws, r1, c1, value, **kwargs)
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = kwargs.get("border", BORDER)
            if kwargs.get("fill"):
                cell.fill = kwargs["fill"]


def _ensure_template() -> Path:
    if TEMPLATE_PATH.is_file():
        return TEMPLATE_PATH

    ASSETS_DIR.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PROFORMA"
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_view.showGridLines = False

    ws.column_dimensions["A"].width = 9
    ws.column_dimensions["B"].width = 38
    ws.column_dimensions["C"].width = 5
    ws.column_dimensions["D"].width = 5
    ws.column_dimensions["E"].width = 11
    ws.column_dimensions["F"].width = 8
    ws.column_dimensions["G"].width = 12

    grey_fill = PatternFill(start_color=GREY, end_color=GREY, fill_type="solid")

    for r in range(1, 5):
        ws.row_dimensions[r].height = 16
    ws.row_dimensions[5].height = 22

    logo = _asset("logo_isitek.png")
    if logo:
        img = XLImage(str(logo))
        img.height = 58
        img.width = 120
        ws.add_image(img, "A1")

    _merge_fill(ws, 5, 1, 5, 7, "PROFORMA", bold=True, size=13, align="center", fill=grey_fill, border=None)

    # Meta gauche + client droite (lignes 6-8)
    for r, label in [(6, "DATE EMISSION:"), (7, "AFFAIRE SUIVIE PAR:"), (8, "REF DEMANDE:")]:
        ws.row_dimensions[r].height = 16
        _set_cell(ws, r, 1, label, bold=True, size=9)

    _set_cell(ws, 6, 5, "A l'attention de:", size=9, align="right")
    _merge_fill(ws, 7, 5, 7, 7, "", align="center", border=None)

    # Tableau contact / phone
    _set_cell(ws, 8, 5, "Contact", bold=True, size=9, border=BORDER)
    _set_cell(ws, 8, 6, "", bold=True, size=9, border=BORDER)
    _merge_fill(ws, 8, 6, 8, 7, "", border=BORDER)
    _set_cell(ws, 9, 5, "Phone", bold=True, size=9, border=BORDER)
    _merge_fill(ws, 9, 6, 9, 7, "", bold=True, border=BORDER)

    ws.row_dimensions[9].height = 16

    # Objet
    ws.row_dimensions[10].height = 18
    _merge_fill(ws, 10, 1, 10, 7, "OBJET DEMANDE:", bold=True, red=True, fill=grey_fill, border=None)

    # En-tête tableau (ligne 11)
    headers = [
        ("Ref", True, False),
        ("DESIGNATION", False, True),
        ("Unit", True, False),
        ("Qté", True, False),
        ("Prix Unit\n(F CFA)", True, False),
        ("REMISE", False, True),
        ("Prix Tot. HT\n(F CFA)", True, False),
    ]
    ws.row_dimensions[ROW_HEADER + 1].height = 28
    for col, (text, italic, bold) in enumerate(headers, start=1):
        _set_cell(ws, ROW_HEADER + 1, col, text, bold=bold, italic=italic, size=8, align="center", fill=grey_fill)

    # Lignes articles vides
    for i in range(ITEM_ROWS):
        row = ROW_ITEMS_START + 1 + i
        ws.row_dimensions[row].height = 16
        for col in range(1, 8):
            _set_cell(ws, row, col, "", border=BORDER)

    # Totaux (colonnes E-F label, G valeur) — comme ze.jpeg
    total_labels = [
        "TOTAL HT BRUT",
        "TOTAL REMISE COMMERCIALE",
        "S/TOTAL HT",
        "REMISE EXCEPTIONNELLE (10%)",
        "TOTAL HT NET",
    ]
    for i, label in enumerate(total_labels):
        row = ROW_TOTALS_START + 1 + i
        ws.row_dimensions[row].height = 16
        fill = grey_fill if "EXCEPTIONNELLE" in label else None
        bord = THICK_TOP if label == "TOTAL HT NET" else BORDER
        for col in range(1, 5):
            _set_cell(ws, row, col, "", border=bord, fill=fill)
        _merge_fill(ws, row, 5, row, 6, label, bold=True, size=8, align="left", fill=fill, border=bord)
        _set_cell(ws, row, 7, "", bold=True, size=8, align="right", fill=fill, border=bord)

    row_svc = ROW_TOTALS_START + 7
    ws.row_dimensions[row_svc].height = 14
    _merge_fill(ws, row_svc, 5, row_svc, 7, "SERVICE COMMERCIAL", bold=True, italic=True, size=9, align="right", border=None)

    row_terms = row_svc + 1
    stamp = _asset("stamp_isitek.png")
    if stamp:
        simg = XLImage(str(stamp))
        simg.height = 72
        simg.width = 155
        ws.add_image(simg, f"A{row_terms}")

    terms_labels = [
        "Validité offre",
        "Delai de livraison",
        "Condition de règlement",
        "Moyen de règlement",
        "Libellé du chèque",
        "Devise",
    ]
    for i, lbl in enumerate(terms_labels):
        r = row_terms + i
        ws.row_dimensions[r].height = 14
        _set_cell(ws, r, 5, f"{lbl} :", bold=True, size=8, border=None)
        _merge_fill(ws, r, 6, r, 7, "", bold=True, size=8, border=None)

    row_brands = row_terms + 8
    brands = _asset("marques_partenaires.png")
    if brands:
        bimg = XLImage(str(brands))
        bimg.width = 520
        bimg.height = 105
        ws.add_image(bimg, f"A{row_brands}")

    row_footer = row_brands + 7
    _merge_fill(
        ws, row_footer, 1, row_footer, 7,
        "ISITEK S.A.R.L au capital de 10.000.000 F CFA - Siège social : Abidjan Cocody Angré - RCCM : CI-ABJ-2017-B-20181",
        size=7, align="center", border=None,
    )
    _merge_fill(
        ws, row_footer + 1, 1, row_footer + 1, 7,
        "BICICI : CI006-01693-010577100067-64 contact@isitek.ci/ TEL: (+225) 25 20 01 19 82 / (+225) 07 59 48 21 84",
        size=7, align="center", border=None,
    )

    wb.save(TEMPLATE_PATH)
    return TEMPLATE_PATH


def _fill_workbook(ws, data: DevisRenderRequest) -> None:
    lignes = data.lignes or []
    brut, remise, sous, rx, net = _compute_totals(
        lignes, data.remise_exceptionnelle_active, data.remise_exceptionnelle_pct
    )

    date_str = data.date_emission
    if not date_str and data.date_devis:
        date_str = data.date_devis.strftime("%d/%m/%Y")
    if not date_str:
        date_str = datetime.now().strftime("%d/%m/%Y")

    # PROFORMA + numéro (rouge)
    num = data.numero_devis or ""
    cell = ws.cell(row=5, column=1)
    if num:
        cell.value = CellRichText(
            TextBlock(InlineFont(rFont=FONT, sz=13, b=True), "PROFORMA   "),
            TextBlock(InlineFont(rFont=FONT, sz=13, b=True, color=RED), num),
        )
    else:
        cell.value = "PROFORMA"
        cell.font = Font(name=FONT, size=13, bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center")

    _set_cell(ws, 6, 2, date_str, bold=True, red=True, border=None)
    _set_cell(ws, 7, 2, data.affaire_suivie_par or "Amadou OUATTARA", bold=True, border=None)
    _set_cell(ws, 8, 2, data.ref_demande or data.client_da or "N/A", bold=True, red=True, border=None)

    _merge_fill(ws, 7, 5, 7, 7, data.client_nom or "", bold=True, align="center", border=BORDER)
    _merge_fill(ws, 8, 6, 8, 7, data.contact or "", bold=True, border=BORDER)
    _merge_fill(ws, 9, 6, 9, 7, data.telephone or data.client_numero_cc or "", bold=True, border=BORDER)

    objet = data.objet_demande or ""
    obj_cell = ws.cell(row=10, column=1)
    obj_cell.value = CellRichText(
        TextBlock(InlineFont(rFont=FONT, sz=9, b=True, color=RED), "OBJET DEMANDE: "),
        TextBlock(InlineFont(rFont=FONT, sz=9, b=True), objet),
    )

    # Articles
    for i in range(ITEM_ROWS):
        row = ROW_ITEMS_START + 1 + i
        for col in range(1, 8):
            ws.cell(row=row, column=col, value="")
        if i < len(lignes):
            l = lignes[i]
            line_brut = l.quantite * l.prix_unitaire_ht
            line_net = line_brut * (1 - l.remise_pourcentage / 100)
            unite = getattr(l, "unite", None) or "U"
            _set_cell(ws, row, 1, l.reference or "", align="center")
            _set_cell(ws, row, 2, (l.designation or "").upper())
            _set_cell(ws, row, 3, unite, align="center")
            _set_cell(ws, row, 4, int(l.quantite) if l.quantite == int(l.quantite) else l.quantite, align="center")
            _set_cell(ws, row, 5, _fmt_num(l.prix_unitaire_ht), bold=True, align="right")
            _set_cell(ws, row, 6, f"{int(l.remise_pourcentage)}%" if l.remise_pourcentage else "", bold=True, align="center")
            _set_cell(ws, row, 7, _fmt_num(line_net), bold=True, align="right")

    totals_vals = [brut, remise, sous, rx if data.remise_exceptionnelle_active else "", net]
    for i, val in enumerate(totals_vals):
        row = ROW_TOTALS_START + 1 + i
        ws.cell(row=row, column=7, value=_fmt_num(val) if val != "" else "")

    rx_label_row = ROW_TOTALS_START + 4
    ws.cell(row=rx_label_row, column=5, value=f"REMISE EXCEPTIONNELLE ({int(data.remise_exceptionnelle_pct)}%)")

    row_terms = ROW_TOTALS_START + 8
    if getattr(data, "condition_reglement", "habituelles") == "acompte":
        cond_label = f"Acompte {getattr(data, 'acompte_pourcentage', 40)}%"
    else:
        cond_label = "Conditions habituelles"
    term_vals = [
        data.validite_offre or "1 mois",
        data.delai_livraison or "1 semaine",
        cond_label,
        data.moyen_reglement or "Chèque/Virement",
        data.libelle_cheque or "ISITEK",
        "Franc CFA (XOF)",
    ]
    for i, val in enumerate(term_vals):
        _merge_fill(ws, row_terms + i, 6, row_terms + i, 7, val, bold=True, size=8, border=None)


def build_devis_excel(data: DevisRenderRequest) -> bytes:
    template = _ensure_template()
    wb = openpyxl.load_workbook(template)
    ws = wb.active
    _fill_workbook(ws, data)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _xlsx_to_pdf_win32(xlsx_path: str, pdf_path: str) -> bool:
    try:
        import win32com.client  # type: ignore

        excel = win32com.client.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        wb = excel.Workbooks.Open(os.path.abspath(xlsx_path))
        wb.ExportAsFixedFormat(0, os.path.abspath(pdf_path))
        wb.Close(False)
        excel.Quit()
        return os.path.isfile(pdf_path)
    except Exception:
        return False


def _xlsx_to_pdf_libreoffice(xlsx_path: str, pdf_path: str) -> bool:
    out_dir = os.path.dirname(pdf_path)
    for cmd in (
        ["soffice", "--headless", "--convert-to", "pdf", "--outdir", out_dir, xlsx_path],
        ["libreoffice", "--headless", "--convert-to", "pdf", "--outdir", out_dir, xlsx_path],
    ):
        try:
            subprocess.run(cmd, check=True, capture_output=True, timeout=120)
            base = os.path.splitext(os.path.basename(xlsx_path))[0] + ".pdf"
            generated = os.path.join(out_dir, base)
            if os.path.isfile(generated) and generated != pdf_path:
                os.replace(generated, pdf_path)
            if os.path.isfile(pdf_path):
                return True
        except Exception:
            continue
    return False


def xlsx_bytes_to_pdf(xlsx_bytes: bytes) -> bytes:
    with tempfile.TemporaryDirectory() as tmp:
        xlsx_path = os.path.join(tmp, "devis.xlsx")
        pdf_path = os.path.join(tmp, "devis.pdf")
        with open(xlsx_path, "wb") as f:
            f.write(xlsx_bytes)
        if _xlsx_to_pdf_win32(xlsx_path, pdf_path) or _xlsx_to_pdf_libreoffice(xlsx_path, pdf_path):
            with open(pdf_path, "rb") as f:
                return f.read()
    raise RuntimeError(
        "Conversion Excel→PDF impossible. Installez Microsoft Excel ou LibreOffice sur le serveur API."
    )


def build_devis_pdf(data: DevisRenderRequest) -> bytes:
    xlsx = build_devis_excel(data)
    try:
        return xlsx_bytes_to_pdf(xlsx)
    except RuntimeError:
        # Fallback : retourner l'Excel si PDF indisponible (évite crash total)
        raise


def devis_model_to_render(devis: DevisProforma) -> DevisRenderRequest:
    lignes = [ProduitLigne(**l) for l in devis.get_lignes()]
    return DevisRenderRequest(
        numero_devis=devis.numero_devis,
        contact=devis.contact,
        client_nom=devis.client_nom,
        client_numero_cc=devis.client_numero_cc,
        client_da=devis.client_da,
        telephone=devis.telephone or devis.client_numero_cc,
        ref_demande=devis.ref_demande or devis.client_da,
        objet_demande=devis.objet_demande or "",
        validite_offre=devis.validite_offre,
        delai_livraison=devis.delai_livraison,
        moyen_reglement=devis.moyen_reglement,
        libelle_cheque=devis.libelle_cheque,
        remise_exceptionnelle_active=bool(devis.remise_exceptionnelle_active),
        remise_exceptionnelle_pct=float(devis.remise_exceptionnelle_pct or 10),
        condition_reglement=getattr(devis, "condition_reglement", None) or "habituelles",
        acompte_pourcentage=int(devis.acompte_pourcentage or 40),
        lignes=lignes,
        date_devis=devis.date_devis,
        affaire_suivie_par=devis.affaire_suivie_par or "Amadou OUATTARA",
    )
